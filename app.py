import os
import cv2
import queue
import threading
import time
import math
from datetime import datetime
import openpyxl
from tkinter import filedialog, messagebox
import customtkinter as ctk
from PIL import Image

# Importamos nuestro detector modular
from detector import (
    PoseDetector, 
    calcular_flexion_tronco, 
    calcular_flexion_cuello,
    obtener_landmarks_analisis, 
    dibujar_analisis_completo
)

# Configuración estética global
ctk.set_appearance_mode("dark")  # Modo oscuro por defecto
ctk.set_default_color_theme("blue")  # Tema de color azul

# Mapeo de colores amigables (HEX para CustomTkinter)
COLOR_PALETTE = {
    "Celeste": {"hex": "#00bfff"},
    "Verde Neón": {"hex": "#00ff64"},
    "Rojo Coral": {"hex": "#ff5050"},
    "Naranja": {"hex": "#ff8c00"},
    "Amarillo": {"hex": "#ffe600"}
}

# Colores predeterminados de análisis (BGR para OpenCV)
MORADO = (128, 0, 128)
AZUL_OSCURO = (139, 0, 0)
BLANCO = (255, 255, 255)
VERDE = (0, 255, 0)

class VideoProcessorThread(threading.Thread):
    """Hilo secundario para procesar video sin congelar la interfaz."""
    def __init__(self, file_path, detector, result_queue):
        super().__init__()
        self.file_path = file_path
        self.detector = detector
        self.result_queue = result_queue
        
        self.running = True
        self.paused = False
        
        self.cap = cv2.VideoCapture(file_path)
        self.fps = self.cap.get(cv2.CAP_PROP_FPS)
        if self.fps <= 0 or math.isnan(self.fps):
            self.fps = 30.0
            
        self.width = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        self.height = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        
        # Limitador de FPS
        self.frame_delay = 1.0 / self.fps

    def run(self):
        frame_idx = 0
        while self.running:
            if self.paused:
                time.sleep(0.1)
                continue
                
            start_time = time.time()
            ret, frame = self.cap.read()
            if not ret:
                # Fin del video
                self.result_queue.put(("EOF", None, None, None))
                break
                
            # Procesar frame con MediaPipe
            landmarks = self.detector.procesar_frame(frame)
            
            # Colocar en la cola de resultados
            if self.result_queue.full():
                try:
                    self.result_queue.get_nowait()
                except queue.Empty:
                    pass
                    
            self.result_queue.put(("FRAME", frame, landmarks, frame_idx))
            frame_idx += 1
            
            # Calcular tiempo de procesamiento y dormir el resto
            elapsed = time.time() - start_time
            sleep_time = max(0.001, self.frame_delay - elapsed)
            time.sleep(sleep_time)
            
        self.cap.release()

    def stop(self):
        self.running = False


class AngleDetectorApp(ctk.CTk):
    """Aplicación principal de escritorio."""
    def __init__(self):
        super().__init__()

        # Configuración de ventana
        self.title("Sistema de Análisis de Flexión del Tronco")
        self.geometry("1200x750")
        self.minsize(1050, 650)
        
        # Inicializar detector
        try:
            self.detector = PoseDetector(confidence_threshold=0.5)
        except Exception as e:
            messagebox.showerror(
                "Error de Inicialización", 
                f"No se pudo inicializar MediaPipe:\n{e}\n\nAsegúrate de tener conexión a Internet para descargar el modelo la primera vez."
            )
            self.destroy()
            return
            
        # Variables de estado
        self.current_filepath = None
        self.file_type = None  # 'image' o 'video'
        self.video_thread = None
        self.result_queue = queue.Queue(maxsize=15)
        
        # Estadísticas de sesión
        
        # Seguimiento de permanencia en el mismo ángulo por segmentos
        self.active_segment = None
        self.video_segments = []
        self.tiempo_permanencia = 0.0

        # Configurar la UI
        self.crear_interfaz()
        
        # Iniciar bucle de lectura de cola para videos
        self.poll_results()

    def crear_interfaz(self):
        # Configuración de Grid Layout principal (1 fila, 2 columnas)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # ---------------------------------------------------------------------
        # COLUMNA IZQUIERDA: SIDEBAR DE CONTROL
        # ---------------------------------------------------------------------
        self.sidebar = ctk.CTkFrame(self, width=320, corner_radius=0)
        self.sidebar.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)
        self.sidebar.grid_rowconfigure(11, weight=1)  # Spacer row weight

        # Título principal
        self.lbl_titulo = ctk.CTkLabel(
            self.sidebar, 
            text="DETECTOR DE ÁNGULOS", 
            font=ctk.CTkFont(family="Helvetica", size=20, weight="bold"),
            text_color="#00bfff"
        )
        self.lbl_titulo.grid(row=0, column=0, padx=20, pady=(25, 5), sticky="w")
        
        self.lbl_subtitulo = ctk.CTkLabel(
            self.sidebar, 
            text="Flexión de Tronco en Posturas", 
            font=ctk.CTkFont(family="Helvetica", size=12, slant="italic"),
            text_color="gray"
        )
        self.lbl_subtitulo.grid(row=1, column=0, padx=20, pady=(0, 20), sticky="w")

        # --- SECCIÓN: CARGA DE ARCHIVO ---
        self.frm_archivo = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        self.frm_archivo.grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        
        self.btn_seleccionar = ctk.CTkButton(
            self.frm_archivo, 
            text="Seleccionar Imagen / Video", 
            command=self.seleccionar_archivo,
            font=ctk.CTkFont(size=13, weight="bold"),
            height=38
        )
        self.btn_seleccionar.pack(fill="x", pady=(0, 5))
        
        self.lbl_archivo = ctk.CTkLabel(
            self.frm_archivo, 
            text="Ningún archivo cargado", 
            wraplength=260,
            text_color="gray",
            font=ctk.CTkFont(size=11)
        )
        self.lbl_archivo.pack(fill="x")

        # Separador visual
        self.separador1 = ctk.CTkFrame(self.sidebar, height=2, fg_color="#333333")
        self.separador1.grid(row=3, column=0, padx=20, pady=15, sticky="ew")

        # --- SECCIÓN: CONFIGURACIÓN DE ANÁLISIS ---
        self.lbl_seccion_conf = ctk.CTkLabel(
            self.sidebar, 
            text="Configuración de Análisis", 
            font=ctk.CTkFont(size=14, weight="bold")
        )
        self.lbl_seccion_conf.grid(row=4, column=0, padx=20, pady=(0, 10), sticky="w")

        # Contenedor para la configuración de análisis
        self.frm_analisis = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        self.frm_analisis.grid(row=5, column=0, padx=20, pady=0, sticky="ew")

        # Lado del cuerpo
        self.lbl_lado = ctk.CTkLabel(self.frm_analisis, text="Lado del cuerpo a medir:")
        self.lbl_lado.pack(anchor="w", padx=5, pady=(0, 2))
        self.opt_lado = ctk.CTkOptionMenu(
            self.frm_analisis, 
            values=["Auto", "Izquierdo", "Derecho"],
            command=self.actualizar_analisis_imagen
        )
        self.opt_lado.pack(fill="x", padx=5, pady=(0, 15))
        self.opt_lado.set("Auto")

        # Confianza del detector
        self.lbl_confianza = ctk.CTkLabel(self.frm_analisis, text="Confianza mínima: 0.50")
        self.lbl_confianza.pack(anchor="w", padx=5, pady=(5, 2))
        self.sld_confianza = ctk.CTkSlider(
            self.frm_analisis, 
            from_=0.1, 
            to=0.9, 
            number_of_steps=16,
            command=self.on_confianza_change
        )
        self.sld_confianza.pack(fill="x", padx=5, pady=(0, 10))
        self.sld_confianza.set(0.5)

        # Separador visual para registro
        self.separador_excel = ctk.CTkFrame(self.sidebar, height=2, fg_color="#333333")
        self.separador_excel.grid(row=6, column=0, padx=20, pady=15, sticky="ew")

        # --- SECCIÓN: REGISTRO DE DATOS ---
        self.lbl_seccion_excel = ctk.CTkLabel(
            self.sidebar, 
            text="Registro de Datos (Excel)", 
            font=ctk.CTkFont(size=14, weight="bold")
        )
        self.lbl_seccion_excel.grid(row=7, column=0, padx=20, pady=(0, 10), sticky="w")

        self.frm_excel = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        self.frm_excel.grid(row=8, column=0, padx=20, pady=0, sticky="ew")

        # Campo nombre
        self.lbl_nombre = ctk.CTkLabel(self.frm_excel, text="Nombre de la Persona:")
        self.lbl_nombre.pack(anchor="w", padx=5, pady=(0, 2))
        self.ent_nombre = ctk.CTkEntry(
            self.frm_excel, 
            placeholder_text="Ej. Juan Pérez"
        )
        self.ent_nombre.pack(fill="x", padx=5, pady=(0, 10))

        # Botón registrar
        self.btn_registrar_excel = ctk.CTkButton(
            self.frm_excel, 
            text="Registrar en Excel", 
            command=self.registrar_en_excel,
            fg_color="#00bfff", 
            hover_color="#008b8b",
            font=ctk.CTkFont(weight="bold")
        )
        self.btn_registrar_excel.pack(fill="x", padx=5, pady=(0, 5))

        # --- SECCIÓN: CONTROLES DE REPRODUCCIÓN (PARA VIDEO) ---
        self.frm_controles_video = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        self.frm_controles_video.grid(row=9, column=0, padx=20, pady=15, sticky="ew")
        
        self.btn_play = ctk.CTkButton(
            self.frm_controles_video, 
            text="Reproducir", 
            fg_color="#10b981", 
            hover_color="#059669",
            command=self.play_video,
            state="disabled",
            font=ctk.CTkFont(weight="bold")
        )
        self.btn_play.pack(side="left", expand=True, padx=(0, 5), fill="x")
        
        self.btn_pause = ctk.CTkButton(
            self.frm_controles_video, 
            text="Pausar", 
            fg_color="#f59e0b", 
            hover_color="#d97706",
            command=self.pause_video,
            state="disabled",
            font=ctk.CTkFont(weight="bold")
        )
        self.btn_pause.pack(side="left", expand=True, padx=(5, 0), fill="x")

        # Botón guardar captura
        self.btn_guardar = ctk.CTkButton(
            self.sidebar, 
            text="Guardar Captura de Imagen", 
            fg_color="#4b5563",
            hover_color="#374151",
            command=self.guardar_captura,
            state="disabled",
            font=ctk.CTkFont(weight="bold"),
            height=35
        )
        self.btn_guardar.grid(row=10, column=0, padx=20, pady=(0, 20), sticky="ew")

        # ---------------------------------------------------------------------
        # COLUMNA DERECHA: VISUALIZADOR Y MÉTRICAS
        # ---------------------------------------------------------------------
        self.visualizer_container = ctk.CTkFrame(self, fg_color="#101010")
        self.visualizer_container.grid(row=0, column=1, sticky="nsew", padx=15, pady=15)
        self.visualizer_container.grid_rowconfigure(0, weight=1)
        self.visualizer_container.grid_columnconfigure(0, weight=1)

        # Panel de visualización de imagen
        self.lbl_viewer = ctk.CTkLabel(
            self.visualizer_container, 
            text="Carga un archivo para comenzar\n\nSoporta: .jpg, .png, .mp4, .avi, .mov",
            font=ctk.CTkFont(family="Segoe UI", size=16),
            text_color="gray"
        )
        self.lbl_viewer.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)

        # --- PANEL DE ESTADÍSTICAS ---
        self.frm_dashboard = ctk.CTkFrame(self.visualizer_container, fg_color="#181818", height=140)
        self.frm_dashboard.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 20))
        self.frm_dashboard.grid_columnconfigure((0, 1, 2, 3), weight=1)

        # Tarjeta 1: Ángulo del Tronco
        self.card_tronco = ctk.CTkFrame(self.frm_dashboard, fg_color="#222222", corner_radius=8)
        self.card_tronco.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        
        lbl_t1 = ctk.CTkLabel(self.card_tronco, text="ÁNGULO TRONCO", font=ctk.CTkFont(size=10, weight="bold"), text_color="gray")
        lbl_t1.pack(pady=(8, 0))
        self.lbl_val_tronco_act = ctk.CTkLabel(self.card_tronco, text="--°", font=ctk.CTkFont(size=28, weight="bold"), text_color="white")
        self.lbl_val_tronco_act.pack(pady=(2, 8))

        # Tarjeta 2: Ángulo del Cuello
        self.card_cuello = ctk.CTkFrame(self.frm_dashboard, fg_color="#222222", corner_radius=8)
        self.card_cuello.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
        
        lbl_t2 = ctk.CTkLabel(self.card_cuello, text="ÁNGULO CUELLO", font=ctk.CTkFont(size=10, weight="bold"), text_color="gray")
        lbl_t2.pack(pady=(8, 0))
        self.lbl_val_cuello_act = ctk.CTkLabel(self.card_cuello, text="--°", font=ctk.CTkFont(size=28, weight="bold"), text_color="white")
        self.lbl_val_cuello_act.pack(pady=(2, 8))

        # Tarjeta 3: Lado Detectado
        self.card_lado = ctk.CTkFrame(self.frm_dashboard, fg_color="#222222", corner_radius=8)
        self.card_lado.grid(row=0, column=2, padx=10, pady=10, sticky="nsew")
        
        lbl_t3 = ctk.CTkLabel(self.card_lado, text="LADO LEÍDO", font=ctk.CTkFont(size=10, weight="bold"), text_color="gray")
        lbl_t3.pack(pady=(8, 0))
        self.lbl_val_lado = ctk.CTkLabel(self.card_lado, text="--", font=ctk.CTkFont(size=22, weight="bold"), text_color="white")
        self.lbl_val_lado.pack(pady=(8, 8))

        # Tarjeta 4: Tiempo en Mismo Ángulo
        self.card_tiempo = ctk.CTkFrame(self.frm_dashboard, fg_color="#222222", corner_radius=8)
        self.card_tiempo.grid(row=0, column=3, padx=10, pady=10, sticky="nsew")
        
        lbl_t4 = ctk.CTkLabel(self.card_tiempo, text="TIEMPO ESTABLE", font=ctk.CTkFont(size=10, weight="bold"), text_color="gray")
        lbl_t4.pack(pady=(8, 0))
        self.lbl_val_tiempo = ctk.CTkLabel(self.card_tiempo, text="0.0s", font=ctk.CTkFont(size=22, weight="bold"), text_color="white")
        self.lbl_val_tiempo.pack(pady=(8, 8))

    # --- MANEJADORES DE EVENTOS DE CONFIGURACIÓN ---
    def on_confianza_change(self, val):
        self.lbl_confianza.configure(text=f"Confianza mínima: {val:.2f}")
        # Re-inicializar detector con nueva confianza
        try:
            self.detector.close()
            self.detector = PoseDetector(confidence_threshold=val)
            # Re-detectar pose en el frame actual si está detenido/pausado
            is_playing = self.video_thread and self.video_thread.is_alive() and not self.video_thread.paused
            if self.raw_current_frame is not None and not is_playing:
                self.current_landmarks = self.detector.procesar_frame(self.raw_current_frame)
            self.actualizar_analisis_imagen()
        except Exception as e:
            print(f"Error al cambiar confianza: {e}")



    # --- MANEJADORES DE ARCHIVOS ---
    def seleccionar_archivo(self):
        # Detener cualquier reproducción previa
        self.stop_video_thread()
        self.active_segment = None
        self.video_segments = []
        self.tiempo_permanencia = 0.0
        
        file_path = filedialog.askopenfilename(
            title="Seleccionar Imagen o Video",
            filetypes=[
                ("Archivos multimedia", "*.jpg *.png *.jpeg *.mp4 *.avi *.mov"),
                ("Imágenes", "*.jpg *.png *.jpeg"),
                ("Videos", "*.mp4 *.avi *.mov")
            ]
        )
        if not file_path:
            return
            
        self.current_filepath = file_path
        file_name = os.path.basename(file_path)
        self.lbl_archivo.configure(text=file_name, text_color="white")
        
        # Determinar tipo de archivo por extensión
        ext = os.path.splitext(file_path)[1].lower()
        if ext in [".jpg", ".jpeg", ".png"]:
            self.file_type = "image"
            self.btn_play.configure(state="disabled")
            self.btn_pause.configure(state="disabled")
            self.btn_guardar.configure(state="normal")
            
            # Cargar imagen
            
            frame = cv2.imread(file_path)
            if frame is not None:
                self.raw_current_frame = frame
                self.current_landmarks = self.detector.procesar_frame(frame)
                self.actualizar_analisis_imagen()
            else:
                self.lbl_viewer.configure(text="Error al leer la imagen seleccionada.")
        else:
            self.file_type = "video"
            self.btn_play.configure(state="normal")
            self.btn_pause.configure(state="disabled")
            self.btn_guardar.configure(state="disabled")
            
            # Mostrar primer frame como previsualización
            self.previsualizar_primer_frame()

    def previsualizar_primer_frame(self):
        """Muestra el primer frame del video como vista previa estática con detección de pose."""
        if not self.current_filepath:
            return
            
        cap = cv2.VideoCapture(self.current_filepath)
        ret, frame = cap.read()
        cap.release()
        
        if ret:
            self.raw_current_frame = frame
            # Detectar pose en el primer frame para visualización inmediata
            self.current_landmarks = self.detector.procesar_frame(frame)
            self.actualizar_analisis_imagen()
            
            pass
        else:
            self.lbl_viewer.configure(text="Error al cargar la previsualización del video.")

    def actualizar_analisis_imagen(self, *args, frame_idx=None):
        """Dibuja sobre el frame actual usando los landmarks guardados y actualiza la GUI."""
        if self.raw_current_frame is None:
            return
            
        frame_a_dibujar = self.raw_current_frame.copy()
        h, w, _ = frame_a_dibujar.shape
        lado_config = self.opt_lado.get()
        
        if self.current_landmarks:
            p_hombro, p_cadera, p_oreja, lado_usado = obtener_landmarks_analisis(self.current_landmarks, lado_config, w, h)
            angulo_tronco = calcular_flexion_tronco(p_hombro, p_cadera)
            angulo_cuello = calcular_flexion_cuello(p_cadera, p_hombro, p_oreja)
            
            # Calcular permanencia de postura en el mismo ángulo
            angulo_tronco_int = int(round(angulo_tronco))
            angulo_cuello_int = int(round(angulo_cuello))
            
            # Calcular tiempo de video/reproducción
            if self.file_type == "video" and frame_idx is not None and self.video_thread is not None:
                t = frame_idx / self.video_thread.fps
            else:
                t = 0.0
                
            # Lógica de segmentos de postura
            if self.file_type == "video" and frame_idx is not None:
                if self.active_segment is None:
                    self.active_segment = {
                        "tronco": angulo_tronco_int,
                        "cuello": angulo_cuello_int,
                        "lado": lado_usado,
                        "t_inicio": t,
                        "t_fin": t
                    }
                    self.tiempo_permanencia = 0.0
                else:
                    if (self.active_segment["tronco"] == angulo_tronco_int and 
                        self.active_segment["cuello"] == angulo_cuello_int and 
                        self.active_segment["lado"] == lado_usado):
                        self.active_segment["t_fin"] = t
                        self.tiempo_permanencia = t - self.active_segment["t_inicio"]
                    else:
                        self.active_segment["duracion"] = self.active_segment["t_fin"] - self.active_segment["t_inicio"]
                        self.video_segments.append(self.active_segment)
                        self.active_segment = {
                            "tronco": angulo_tronco_int,
                            "cuello": angulo_cuello_int,
                            "lado": lado_usado,
                            "t_inicio": self.active_segment["t_fin"],
                            "t_fin": t
                        }
                        self.tiempo_permanencia = t - self.active_segment["t_inicio"]
            else:
                self.tiempo_permanencia = 0.0

            dibujo_config = {
                "color_vertical": MORADO,           # Morado
                "color_tronco": MORADO,             # Morado
                "color_cuello": MORADO,             # Morado
                "color_arco": VERDE,                # Verde
                "color_puntos": AZUL_OSCURO,        # Azul Oscuro
                "color_texto": BLANCO,              # Blanco
                "grosor_lineas": 3,
                "grosor_tronco": 5,
                "grosor_cuello": 4,
                "grosor_borde_arco": 2,
                "radio_arco": 40,
                "radio_arco_cuello": 30,            # 77% de 40 es 30.8, redondeado a 30
                "radio_ear": 6,                     # Círculo pequeño para la oreja
                "radio_hombro": 22,                 # Tamaño suficiente para que quepa el texto
                "radio_cadera": 22,                 # Tamaño suficiente para que quepa el texto
                "transparencia_arco": 0.28,         # Opacidad de 28%
                "dibujar_texto": True
            }
            
            dibujar_analisis_completo(frame_a_dibujar, p_cadera, p_hombro, p_oreja, angulo_tronco, angulo_cuello, config=dibujo_config)
            
            # Guardar frame procesado actual en memoria
            self.current_processed_frame = frame_a_dibujar
            
            # Actualizar métricas
            self.actualizar_metricas_ui(angulo_tronco, angulo_cuello, lado_usado)
        else:
            self.current_processed_frame = frame_a_dibujar
            self.lbl_val_tronco_act.configure(text="--°", text_color="white")
            self.lbl_val_cuello_act.configure(text="--°", text_color="white")
            self.lbl_val_lado.configure(text="--")
            self.lbl_val_tiempo.configure(text="0.0s")
            
            # Lógica de segmentos para frame sin pose
            if self.file_type == "video" and frame_idx is not None and self.video_thread is not None:
                t = frame_idx / self.video_thread.fps
                if self.active_segment is None:
                    self.active_segment = {
                        "tronco": None,
                        "cuello": None,
                        "lado": None,
                        "t_inicio": t,
                        "t_fin": t
                    }
                    self.tiempo_permanencia = 0.0
                else:
                    if (self.active_segment["tronco"] is None and 
                        self.active_segment["cuello"] is None and 
                        self.active_segment["lado"] is None):
                        self.active_segment["t_fin"] = t
                        self.tiempo_permanencia = t - self.active_segment["t_inicio"]
                    else:
                        self.active_segment["duracion"] = self.active_segment["t_fin"] - self.active_segment["t_inicio"]
                        self.video_segments.append(self.active_segment)
                        self.active_segment = {
                            "tronco": None,
                            "cuello": None,
                            "lado": None,
                            "t_inicio": self.active_segment["t_fin"],
                            "t_fin": t
                        }
                        self.tiempo_permanencia = t - self.active_segment["t_inicio"]
            
        self.mostrar_frame_en_viewer(self.current_processed_frame)

    # --- MÉTODOS DE REPRODUCCIÓN DE VIDEO ---
    def play_video(self):
        if self.file_type != "video" or not self.current_filepath:
            return
            
        # Si el hilo ya existe y está pausado, lo reanudamos
        if self.video_thread and self.video_thread.is_alive():
            if self.video_thread.paused:
                self.video_thread.paused = False
                self.btn_play.configure(state="disabled")
                self.btn_pause.configure(state="normal")
                self.btn_guardar.configure(state="disabled")
                self.sld_confianza.configure(state="disabled")
            return
            
        # Iniciar reproducción
        
        # Limpiar cola
        while not self.result_queue.empty():
            try:
                self.result_queue.get_nowait()
            except queue.Empty:
                break
                
        # Crear y arrancar hilo procesador
        self.video_thread = VideoProcessorThread(
            file_path=self.current_filepath,
            detector=self.detector,
            result_queue=self.result_queue
        )
        self.video_thread.start()
        
        self.btn_play.configure(state="disabled")
        self.btn_pause.configure(state="normal")
        self.btn_guardar.configure(state="disabled")
        self.sld_confianza.configure(state="disabled")

    def pause_video(self):
        if self.video_thread and self.video_thread.is_alive():
            self.video_thread.paused = True
            self.btn_play.configure(state="normal", text="Reanudar")
            self.btn_pause.configure(state="disabled")
            self.btn_guardar.configure(state="normal")
            self.sld_confianza.configure(state="normal")

    def stop_video_thread(self):
        """Detiene de forma segura el hilo de reproducción de video."""
        if self.video_thread:
            self.video_thread.stop()
            self.video_thread.join(timeout=1.0)
            self.video_thread = None
        self.btn_play.configure(text="Reproducir", state="disabled")
        self.btn_pause.configure(state="disabled")
        self.btn_guardar.configure(state="disabled")
        self.sld_confianza.configure(state="normal")

    def poll_results(self):
        """Monitorea la cola de resultados en el hilo principal para actualizar la GUI."""
        try:
            # Intentar vaciar tantos frames listos como haya (para mantenerse al día)
            while True:
                item = self.result_queue.get_nowait()
                status = item[0]
                
                if status == "EOF":
                    # El video terminó
                    self.stop_video_thread()
                    if self.current_filepath:
                        self.btn_play.configure(state="normal", text="Reproducir de nuevo")
                    break
                    
                if status == "FRAME":
                    _, frame, landmarks, frame_idx = item
                    if frame is not None:
                        self.raw_current_frame = frame
                        self.current_landmarks = landmarks
                        self.actualizar_analisis_imagen(frame_idx=frame_idx)
                        
        except queue.Empty:
            pass
            
        # Programar la siguiente verificación en 15ms (~60 FPS de sondeo)
        self.after(15, self.poll_results)

    # --- MÉTODOS AUXILIARES ---

    def actualizar_metricas_ui(self, angulo_tronco, angulo_cuello, lado):
        """Actualiza las tarjetas de datos con el ángulo y color correspondiente."""
        self.lbl_val_tronco_act.configure(text=f"{int(round(angulo_tronco))}°")
        self.lbl_val_cuello_act.configure(text=f"{int(round(angulo_cuello))}°")
        self.lbl_val_lado.configure(text=str(lado))
        
        # Lógica de color según inclinación de la espalda
        if angulo_tronco < 15.0:
            color_tronco = COLOR_PALETTE["Verde Neón"]["hex"]
        elif angulo_tronco < 40.0:
            color_tronco = COLOR_PALETTE["Naranja"]["hex"]
        else:
            color_tronco = COLOR_PALETTE["Rojo Coral"]["hex"]
            
        self.lbl_val_tronco_act.configure(text_color=color_tronco)

        # Lógica de color según inclinación del cuello
        if angulo_cuello < 10.0:
            color_cuello = COLOR_PALETTE["Verde Neón"]["hex"]
        elif angulo_cuello < 20.0:
            color_cuello = COLOR_PALETTE["Naranja"]["hex"]
        else:
            color_cuello = COLOR_PALETTE["Rojo Coral"]["hex"]
            
        self.lbl_val_cuello_act.configure(text_color=color_cuello)
        # Mostrar tiempo de permanencia en el mismo ángulo con 5 decimales
        self.lbl_val_tiempo.configure(text=f"{self.tiempo_permanencia:.5f}s")

    def mostrar_frame_en_viewer(self, frame):
        """Adapta la imagen OpenCV BGR para mostrarla centrada en el visor de CustomTkinter."""
        if frame is None:
            return
            
        # Convertir BGR a RGB
        rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        pil_image = Image.fromarray(rgb_image)
        
        # Obtener dimensiones del visor
        view_w = self.lbl_viewer.winfo_width()
        view_h = self.lbl_viewer.winfo_height()
        
        # Si el tamaño es 1x1 (antes de renderizar la ventana por primera vez), usar un valor por defecto
        if view_w <= 1 or view_h <= 1:
            view_w = 800
            view_h = 500
            
        # Escalar manteniendo la relación de aspecto
        img_w, img_h = pil_image.size
        aspect_ratio = img_w / img_h
        
        if view_w / view_h > aspect_ratio:
            new_h = view_h - 10
            new_w = int(new_h * aspect_ratio)
        else:
            new_w = view_w - 10
            new_h = int(new_w / aspect_ratio)
            
        # Asegurarse de que las dimensiones sean válidas
        new_w = max(10, new_w)
        new_h = max(10, new_h)
        
        # Convertir a CTkImage usando la imagen en resolución original
        # para que CustomTkinter mantenga la nitidez (HD) al aplicar escala DPI
        ctk_img = ctk.CTkImage(
            light_image=pil_image,
            dark_image=pil_image,
            size=(new_w, new_h)
        )
        
        # Actualizar widget
        self.lbl_viewer.configure(image=ctk_img, text="")
        self.lbl_viewer.image = ctk_img  # Mantener referencia

    def guardar_captura(self):
        """Guarda la imagen procesada actual en disco."""
        if self.current_processed_frame is None:
            return
            
        save_path = filedialog.asksaveasfilename(
            title="Guardar Captura",
            defaultextension=".png",
            filetypes=[("Imagen PNG", "*.png"), ("Imagen JPEG", "*.jpg")]
        )
        
        if save_path:
            try:
                cv2.imwrite(save_path, self.current_processed_frame)
                messagebox.showinfo("Éxito", f"Captura guardada correctamente en:\n{save_path}")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar la imagen:\n{e}")

    def registrar_en_excel(self):
        """Registra los ángulos actuales y el nombre de la persona en un archivo Excel."""
        nombre = self.ent_nombre.get().strip()
        if not nombre:
            messagebox.showwarning("Campo Vacío", "Por favor, ingrese el nombre de la persona para realizar el registro.")
            return

        if self.current_landmarks is None or self.raw_current_frame is None:
            messagebox.showwarning("Sin Detección", "No hay datos de pose detectados para registrar. Asegúrese de cargar un archivo con una persona visible.")
            return

        # Para videos, verificar que tengamos segmentos procesados para guardar
        if self.file_type == "video" and not self.video_segments and self.active_segment is None:
            messagebox.showwarning("Sin Datos de Video", "No hay datos de video acumulados. Inicie la reproducción del video para analizar y medir los ángulos antes de registrar.")
            return

        # Calcular los valores de la postura actual en pantalla
        h, w, _ = self.raw_current_frame.shape
        lado_config = self.opt_lado.get()
        p_hombro, p_cadera, p_oreja, lado_usado = obtener_landmarks_analisis(self.current_landmarks, lado_config, w, h)
        angulo_tronco = calcular_flexion_tronco(p_hombro, p_cadera)
        angulo_cuello = calcular_flexion_cuello(p_cadera, p_hombro, p_oreja)

        # Obtener nombre del archivo de origen
        archivo_origen = os.path.basename(self.current_filepath) if self.current_filepath else "Previsualización / Imagen"

        # Fecha y hora actual
        fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Nombre del archivo Excel a guardar
        excel_path = "registro_posturas.xlsx"

        try:
            if os.path.exists(excel_path):
                # Abrir archivo existente
                workbook = openpyxl.load_workbook(excel_path)
                sheet = workbook.active
                
                # Si la hoja está totalmente vacía, añadimos cabeceras
                if sheet.max_row == 0 or (sheet.max_row == 1 and all(cell.value is None for cell in sheet[1])):
                    sheet.delete_rows(1, sheet.max_row)
                    sheet.append([
                        "Fecha y Hora",
                        "Nombre de la Persona",
                        "Ángulo Tronco (Grados)",
                        "Tiempo Tronco (Segundos)",
                        "Ángulo Cuello (Grados)",
                        "Tiempo Cuello (Segundos)",
                        "Lado del Cuerpo Medido",
                        "Archivo de Origen"
                    ])
                else:
                    # Limpiar columnas viejas si existen
                    headers = [cell.value for cell in sheet[1]]
                    for col in ["Estado del Cuello", "Estado del Tronco", "Tiempo de Video (Segundos)", "Tiempo en Mismo Ángulo (Segundos)"]:
                        if col in headers:
                            col_idx = headers.index(col) + 1
                            sheet.delete_cols(col_idx, 1)
                            headers = [cell.value for cell in sheet[1]]
                    
                    # Si falta "Tiempo Tronco (Segundos)", insertarla como columna 4
                    if "Tiempo Tronco (Segundos)" not in headers:
                        sheet.insert_cols(4, 1)
                        headers = [cell.value for cell in sheet[1]]

                    # Si falta "Tiempo Cuello (Segundos)", insertarla como columna 6
                    if "Tiempo Cuello (Segundos)" not in headers:
                        sheet.insert_cols(6, 1)
                        headers = [cell.value for cell in sheet[1]]
                        
                    # Forzar los nombres exactos en la fila 1 para cabeceras limpias
                    sheet.cell(row=1, column=1, value="Fecha y Hora")
                    sheet.cell(row=1, column=2, value="Nombre de la Persona")
                    sheet.cell(row=1, column=3, value="Ángulo Tronco (Grados)")
                    sheet.cell(row=1, column=4, value="Tiempo Tronco (Segundos)")
                    sheet.cell(row=1, column=5, value="Ángulo Cuello (Grados)")
                    sheet.cell(row=1, column=6, value="Tiempo Cuello (Segundos)")
                    sheet.cell(row=1, column=7, value="Lado del Cuerpo Medido")
                    sheet.cell(row=1, column=8, value="Archivo de Origen")
            else:
                # Crear nuevo archivo y agregar cabeceras
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Registro de Posturas"
                sheet.append([
                    "Fecha y Hora",
                    "Nombre de la Persona",
                    "Ángulo Tronco (Grados)",
                    "Tiempo Tronco (Segundos)",
                    "Ángulo Cuello (Grados)",
                    "Tiempo Cuello (Segundos)",
                    "Lado del Cuerpo Medido",
                    "Archivo de Origen"
                ])

            if self.file_type == "video":
                # Obtener segmentos del video a registrar (incluyendo el activo actual)
                segmentos_a_registrar = list(self.video_segments)
                if self.active_segment is not None:
                    seg_activo_copia = dict(self.active_segment)
                    seg_activo_copia["duracion"] = seg_activo_copia["t_fin"] - seg_activo_copia["t_inicio"]
                    segmentos_a_registrar.append(seg_activo_copia)

                # Escribir todos los segmentos acumulados del video
                num_registros = 0
                for seg in segmentos_a_registrar:
                    if seg["duracion"] <= 0:
                        continue
                    sheet.append([
                        fecha_hora,
                        nombre,
                        seg["tronco"] if seg["tronco"] is not None else "--",
                        round(seg["duracion"], 5),
                        seg["cuello"] if seg["cuello"] is not None else "--",
                        round(seg["duracion"], 5),
                        seg["lado"] if seg["lado"] is not None else "--",
                        archivo_origen
                    ])
                    num_registros += 1
                
                # Limpiar historial tras guardar con éxito
                self.video_segments = []
                if self.active_segment is not None:
                    # Reiniciar el segmento activo a partir del tiempo en que guardamos
                    self.active_segment["t_inicio"] = self.active_segment["t_fin"]
                
                mensaje_exito = f"Se registraron exitosamente {num_registros} segmentos de postura del paciente '{nombre}' en:\n{os.path.abspath(excel_path)}"
            else:
                # Imagen: escribir un solo registro con duración 0.0
                sheet.append([
                    fecha_hora,
                    nombre,
                    int(round(angulo_tronco)),
                    0.0,
                    int(round(angulo_cuello)),
                    0.0,
                    lado_usado,
                    archivo_origen
                ])
                mensaje_exito = f"Datos del paciente '{nombre}' registrados correctamente en:\n{os.path.abspath(excel_path)}"

            # Guardar el libro de trabajo
            workbook.save(excel_path)
            workbook.close()

            # Limpiar la caja de texto tras registrar con éxito
            self.ent_nombre.delete(0, 'end')

            messagebox.showinfo("Éxito", mensaje_exito)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo escribir en el archivo de Excel:\n{e}\n\nAsegúrese de que el archivo no esté abierto en otra aplicación.")

    def on_closing(self):
        """Maneja el cierre seguro de la ventana deteniendo los hilos."""
        self.stop_video_thread()
        try:
            self.detector.close()
        except:
            pass
        self.destroy()


if __name__ == "__main__":
    app = AngleDetectorApp()
    app.protocol("WM_DELETE_WINDOW", app.on_closing)
    app.mainloop()
