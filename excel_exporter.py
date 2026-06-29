import os
import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def registrar_posturas_excel(
    excel_path, 
    nombre, 
    archivo_origen, 
    frames_data,
    alfa,
    beta
):
    """
    Registra datos de postura por frame en un archivo Excel con formato de 14 columnas.
    """
    # Cargar o crear libro de trabajo
    if os.path.exists(excel_path):
        workbook = openpyxl.load_workbook(excel_path)
    else:
        workbook = openpyxl.Workbook()

    sheet_name = "Registro Posturas"
    
    # Asegurar que exista la hoja requerida
    if sheet_name not in workbook.sheetnames:
        if not os.path.exists(excel_path) and len(workbook.sheetnames) == 1 and workbook.active.title == "Sheet":
            sheet = workbook.active
            sheet.title = sheet_name
        else:
            sheet = workbook.create_sheet(sheet_name)
    else:
        sheet = workbook[sheet_name]

    # Eliminar hojas antiguas si existen (para migración limpia del sistema anterior)
    old_sheet_names = ["DATOS TRONCO", "CABEZA", "DATOS CUELLO", "DATOS BRAZO", "Sheet"]
    for name in list(workbook.sheetnames):
        if name in old_sheet_names and name != sheet_name:
            try:
                del workbook[name]
            except Exception:
                pass

    # Cabeceras requeridas por el usuario
    headers = [
        "Nombre de la persona",
        "Lado leído",
        "Ángulo de referencia del tronco (α)",
        "Ángulo del tronco del video",
        "Ángulo del tronco ajustado",
        "Ángulo de referencia de la cabeza (β)",
        "Ángulo de la cabeza del video",
        "Ángulo de la cabeza ajustado",
        "Ángulo del cuello ajustado",
        "Ángulo de brazo",
        "Ángulo de la muñeca",
        "Tiempo de postura",
        "Frames acumulados",
        "Nombre del archivo"
    ]

    # Estilos de formato: letras negras (negrita o normal) y alineación con ajustar texto
    bold_black_font = Font(name="Calibri", size=11, bold=True, color="000000")
    regular_black_font = Font(name="Calibri", size=11, color="000000")
    header_alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Asegurar cabeceras en la hoja
    if sheet.max_row == 0 or (sheet.max_row == 1 and all(cell.value is None for cell in sheet[1])):
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(headers)
    else:
        # Forzar cabeceras exactas en fila 1
        for col_idx, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_idx, value=header)

    # Aplicar formato de cabecera a toda la fila 1 (ajustar texto, negrita y sin fondo)
    for col_idx in range(1, 15):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = bold_black_font
        cell.alignment = header_alignment
        cell.fill = PatternFill(fill_type=None)  # Sin fondo

    alfa_val = int(round(alfa)) if alfa is not None else None
    beta_val = int(round(beta)) if beta is not None else None

    # Escribir cada frame
    for f_data in frames_data:
        t_vid = f_data.get("angulo_tronco")
        c_vid = f_data.get("angulo_cabeza")
        
        # Calcular tronco ajustado (Ángulo del tronco del video - Alfa)
        if t_vid is not None and t_vid != "--" and alfa_val is not None:
            try:
                t_vid_int = int(round(float(t_vid)))
                tronco_ajustado = t_vid_int - alfa_val
            except (ValueError, TypeError):
                tronco_ajustado = "--"
        else:
            tronco_ajustado = "--"

        # Calcular cabeza ajustada (Ángulo de la cabeza del video - Beta)
        if c_vid is not None and c_vid != "--" and beta_val is not None:
            try:
                c_vid_int = int(round(float(c_vid)))
                cabeza_ajustada = c_vid_int - beta_val
            except (ValueError, TypeError):
                cabeza_ajustada = "--"
        else:
            cabeza_ajustada = "--"

        # Calcular cuello ajustado si ambos están calculados
        if cabeza_ajustada != "--" and tronco_ajustado != "--":
            cuello_ajustado = cabeza_ajustada - tronco_ajustado
        else:
            cuello_ajustado = "--"

        sheet.append([
            nombre,
            f_data.get("lado_usado") if f_data.get("lado_usado") is not None else "--",
            alfa_val if alfa_val is not None else "--",
            t_vid if t_vid is not None else "--",
            tronco_ajustado,
            beta_val if beta_val is not None else "--",
            c_vid if c_vid is not None else "--",
            cabeza_ajustada,
            cuello_ajustado,
            f_data.get("angulo_hombro") if f_data.get("angulo_hombro") is not None else "--",
            f_data.get("angulo_muneca") if f_data.get("angulo_muneca") is not None else "--",
            f_data.get("tiempo_postura") if f_data.get("tiempo_postura") is not None else 0.0,
            f_data.get("frames_acumulados") if f_data.get("frames_acumulados") is not None else 0,
            archivo_origen
        ])

        # Aplicar formato de datos a las celdas de la fila agregada (sin fondo; negrita en D, E, G, H, I, J, K -> 4, 5, 7, 8, 9, 10, 11)
        last_row = sheet.max_row
        for col_idx in range(1, 15):
            cell = sheet.cell(row=last_row, column=col_idx)
            cell.fill = PatternFill(fill_type=None)  # Sin fondo
            if col_idx in [4, 5, 7, 8, 9, 10, 11]:
                cell.font = bold_black_font
            else:
                cell.font = regular_black_font

    # Configurar el ancho de todas las columnas (de la A a la O, columnas 1 a 15) a 12
    for col_idx in range(1, 16):
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = 12

    # Configurar altura de filas (fila 1 = 45, demás = 14.4)
    sheet.row_dimensions[1].height = 45
    for r in range(2, sheet.max_row + 1):
        sheet.row_dimensions[r].height = 14.4

    workbook.save(excel_path)
    workbook.close()
    return len(frames_data)
